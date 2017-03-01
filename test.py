"""
EASY TESTCASE

Created on Wed Oct 26 14:22:07 2016

@author: dimiboeckaerts
"""

import openpyxl as oxl
import numpy as np
import random as ran
import datetime as dt
from helpers import unicorn
from helpers import indx
from helpers import buds
from helpers import weight
from helpers import highestcost
from helpers import satdegree
from helpers import theverybest
from helpers import fitness1

file = 'stu_vak_ba2.xlsx'
prof = 'prof_ba.xlsx'
days = 'days_ba.xlsx'
length = 28
n1 = 20
n2 = 15


def dataprep(file, days, prof):
    
    # FILE
    wb = oxl.load_workbook(file)
    lst = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(lst[0])
    test = 0
    
    for j in range(1, sheet.max_column):
        sheetvalue = sheet.cell(row=1, column=j).value
        logic1 = sheetvalue in ['VAK', 'Vak', 'vak', 'CURSUS', 'Cursus', 'cursus', 'CURSUSNAAM', 'Cursusnaam', 'cursusnaam']
        logic2 = sheetvalue in ['EMAIL', 'Email', 'email', 'MAIL', 'Mail', 'mail', 'EMAILADRES', 'Emailadres', 'emailadres']
        if logic1 == True:
            test += 1
            lst1 = []
            for i in range(2, sheet.max_row+1):
                lst1.append(sheet.cell(row=i, column=j).value)
                if (i/sheet.max_row*100) % 10 == 0:
                    print(str(i/sheet.max_row*100) + ' % done (1/4)')
        elif logic2 == True:
            test += 1
            lst2 = []
            for k in range(2, sheet.max_row+1):
                lst2.append(sheet.cell(row=k, column=j).value)
                if (k/sheet.max_row*100) % 10 == 0:
                    print(str(k/sheet.max_row*100) + ' % done (2/4)')
    if test != 2:
        print('Oops, it seems that you have not given me the correct input.\
        Please try again.')
        return
    else:
        vak = lst1
        stu = lst2
        courses = unicorn(vak)
    
    # DAYS
    wb = oxl.load_workbook(days)
    lst = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(lst[0])
    
    if (type(sheet.cell(row=1, column=2).value) != int) and (type(sheet.cell(row=1, column=1).value) != str):
        print('Oops, there is something wrong with your \'days\' file. Please try again.')
        return
            
    for i in range(1, sheet.max_row+1):
        sheetvalue = sheet.cell(row=i, column=2).value
        if sheetvalue > 1:
            for j in range(1,sheetvalue):
                # range to sheetvalue, not sheetvalue+1 because we already have
                # the course one time in the list.
                courses.append(sheet.cell(row=i, column=1).value)
        if (i/sheet.max_row*100) % 10 == 0:
            print(str(i/sheet.max_row*100) + ' % done (3/4)')
        
    courses = sorted(courses)
    
    # PROF
    wb = oxl.load_workbook(prof)
    lst = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(lst[0])
    
    if (type(sheet.cell(row=1, column=2).value) != int) and (type(sheet.cell(row=1, column=1).value) != str):
        print('Oops, there is something wrong with your \'prof\' file. Please try again.')
        return
    else:
        profd = {}
        for i in range(1, sheet.max_row+1):
            lijst = []
            for j in range(2, sheet.max_column+1):
                if sheet.cell(row=i, column=j).value != 0:
                    lijst.append(sheet.cell(row=i, column=j).value)
                    
            profd[sheet.cell(row=i, column=1).value] = lijst

            if (i/sheet.max_row*100) % 10 == 0:
                print(str(i/sheet.max_row*100) + ' % done (4/4)')
                
    print('Data prep done. Now let\'s get scheduling.')
    
    return stu, vak, courses, profd

    
# INITIAL SCHEDULE 
# ----------------------------------------
def thecreator(stu, vak, courses, length, profd, n1):
    now = dt.datetime.now()
    print('Starting process to compute initial generation of timetables.')
    vector = [x for x in range(1,length+1)]
    for item in vector:
        if item%7 == 0:
            vector.remove(item)
            vector.remove(item-1)
    cost = []
    sched = {}
        
    # Get the first ordering
    for item in courses:
        cost.append(satdegree(item, profd))
    order = [courses for (cost, courses) in sorted(zip(cost,courses))][::-1]
    
    for i in range(0, n1):
        sched[i] = [0] * len(courses)
        # Plan first exam for each generation (sat degree) (slow)
        firstexam = order[ran.randint(0,2)]
        # update vector list to only possible dates
        vector2 = [x for x in vector if x not in profd[firstexam]]
        cost = []
        for item in vector2:
            cost.append(theverybest(firstexam, stu, vak, courses, sched[i], length, item))    
        order_period = [vector2 for (cost, vector2) in sorted(zip(cost,vector2))]
        index = indx(courses, firstexam)
        
        # Schedule all orals directly
        for k in range(0, len(index)):
            sched[i][index[k]] = order_period[k] # dict, [][]
    
        # Plan rest of courses for each of the individuals in population
        # While exams have to be scheduled in this individu
        while (np.prod(sched[i]) == 0): #or len(unscheduled) != 0:
            # Update list to only unscheduled courses, then use heuristics
            unscheduled_index = indx(sched[i], 0)
            cost = []
            unscheduled = []
            for item in unscheduled_index:
                unscheduled.append(courses[item])
                cost.append(highestcost(courses[item], stu, vak, courses, sched[i], length, profd))
                # now unscheduled matches with cost
            neworder = [unscheduled for (cost, unscheduled) in sorted(zip(cost,unscheduled))][::-1]

            # Schedule most difficult exams (random between two most difficult)
            if len(unscheduled) == 1:
                exam = unscheduled[0]
            else:
                exam = neworder[ran.randint(0,1)]
            vector2 = [x for x in vector if x not in profd[exam]]
            cost = []
            for item in vector2:
                cost.append(theverybest(exam, stu, vak, courses, sched[i], length, item))    
            order_period = [vector2 for (cost, vector2) in sorted(zip(cost,vector2))]
            index = indx(courses, exam)
            for l in range(0, len(index)):
                sched[i][index[l]] = order_period[l]
            #print(sched[i])
            
        if (i/n1*100) % 5 == 0:
            print(str(i/n1*100) + ' %  of initial timetables done')
    print(dt.datetime.now() - now)
    return vector, sched

stu, vak, courses, profd = dataprep(file, days, prof)
vector, sched = thecreator(stu, vak, courses, length, profd, n1)


# PHASE 1  
# ----------------------------------------
def optimizer1(sched, courses, stu, vak, profd):
    newsched = {}
    test = 0
    for i in range(0, len(sched)):
        ids = ran.sample(range(len(sched)), 10)
        fitness = 1000
        for item in ids:
            newfitness, clashes = fitness1(sched[item], courses, stu, vak)
            print(newfitness)
            if newfitness < fitness: #fitness = # of clashes
                fitness = newfitness
                newschedlst = sched[item]
                newclashes = clashes
        newsched[i] = newschedlst
        
        # reschedule first clash to min cost period
        if len(newclashes) > 0:
            test = 1
            vector2 = [x for x in vector if x not in profd[newclashes[0]]]
            cost = []
            for item in vector2:
                cost.append(theverybest(newclashes[0], stu, vak, courses, newsched[i], length, item))
            order_period = [vector2 for (cost, vector2) in sorted(zip(cost, vector2))]
            index = indx(courses, newclashes[0])
            newsched[i][index[0]] = order_period[0]
    return newsched, test

fail_count_1 = 0
test = 1    
while (test != 0 and fail_count_1 < 50):
    sched, test = optimizer1(sched, courses, stu, vak, profd)
    fail_count_1 += 1
        
if fail_count_1 == 50:
    print('Very difficult schedule... full optimization probably impossible.')

    
# PHASE 2   
# ----------------------------------------
def fitness2(schedlst, courses, stu, vak):
    softcost = 0
    
    for i in range(0, len(schedlst)):
        exam1 = courses[i]
        for j in range(i+1, len(schedlst)):
            exam2 = courses[j]
            if (exam1 != exam2):
                softcost += weight(abs(schedlst[i]-schedlst[j]))*buds(exam1, exam2, stu, vak)
    softcost /= len(unicorn(stu))
    
    return softcost
    
def optimizer2(sched, vector, courses, stu, vak, profd, n2):
    newsched = {}

    for i in range(0, n2):
        ids = ran.sample(range(len(sched)), 10)
        softcost = 1000
        for item in ids:
            newcost = fitness2(sched[item], courses, stu, vak)
            print(newcost)
            
            if newcost < softcost:
                softcost = newcost
                newschedlst = sched[item]
        newsched[i] = newschedlst

        times = ran.randint(1, 3)
        for j in range(0, times):    
            fix1 = ran.randint(0, len(sched[i])-1)
            
            # only feasible mutations
            vector2 = [x for x in vector if x not in profd[courses[fix1]]]
            vector3 = [x for x in vector2 if x not in newsched[i]]
        
            tester = 1
            tries = 0
            while (tester != 0 and tries < 3):
                fix2 = ran.sample(vector3, 1)[0]
                testsched = newsched[i]
                testsched[fix1] = fix2
                testcost = fitness2(testsched, courses, stu, vak)
                if testcost < softcost:
                    tester = 0
                    newsched[i] = testsched
                tries += 1
    return newsched

n_gen = 2
now = dt.datetime.now()
for i in range(0, n_gen):
    sched = optimizer2(sched, vector, courses, stu, vak, profd, n2)
    print(sched)
print(dt.datetime.now() - now)


# OUTPUT   
# ----------------------------------------
startdate = '01/06/17'
thedate = dt.datetime.strptime(startdate, "%d/%m/%y")
final_sched = []
softcost = 10000
    
for i in range(0, len(sched)):
    newcost = fitness2(sched[i], courses, stu, vak)
    if newcost < softcost:
        softcost = newcost
        newschedlst = sched[i]
        
for i in range(0, len(newschedlst)):
    date = thedate + dt.timedelta(days = newschedlst[i]-1)
    final_sched.append(date.strftime('%d/%m'))
    
print('Final schedule:')
for i in range(0, len(final_sched)):
    print(courses[i], ':', final_sched[i])
